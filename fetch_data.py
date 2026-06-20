#!/usr/bin/env python3
"""
SUBHA CAPITAL — Market Data Fetcher V10
Original ROCKERS System — Restored
Nifty 100 Universe (91 stocks)
6-condition ROCKERS scan
Top 2 Long + Top 2 Short for GTT
Only 2 smart filters:
  1. First candle moved >2% → Skip
  2. Gap >5% at open → Skip (results day)
"""

import os,json,math,pyotp,requests,pytz,time,openpyxl
from openpyxl.styles import PatternFill,Font,Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime,timedelta

CLIENT_ID   = os.environ["ANGEL_CLIENT_ID"]
API_KEY     = os.environ["ANGEL_API_KEY"]
TOTP_SECRET = os.environ["ANGEL_TOTP_SECRET"]
PIN         = os.environ["ANGEL_PIN"]
CAPITAL     = 100000
RISK_PCT    = 1.0
MARGIN      = 5
IST         = pytz.timezone("Asia/Kolkata")
BASE        = "https://apiconnect.angelone.in"

# ── NIFTY 100 UNIVERSE (91 stocks) ───────────────────────────────────
STOCKS = {
    "HDFCBANK":   {"token":"1333",  "sec":"BANKING"},
    "ICICIBANK":  {"token":"4963",  "sec":"BANKING"},
    "SBIN":       {"token":"3045",  "sec":"BANKING"},
    "AXISBANK":   {"token":"5900",  "sec":"BANKING"},
    "KOTAKBANK":  {"token":"1922",  "sec":"BANKING"},
    "BANKBARODA": {"token":"4668",  "sec":"BANKING"},
    "INDUSINDBK": {"token":"5258",  "sec":"BANKING"},
    "CANBK":      {"token":"10794", "sec":"BANKING"},
    "BAJFINANCE": {"token":"317",   "sec":"FINANCE"},
    "BAJAJFINSV": {"token":"16675", "sec":"FINANCE"},
    "HDFCLIFE":   {"token":"467",   "sec":"FINANCE"},
    "SBILIFE":    {"token":"21808", "sec":"FINANCE"},
    "ICICIPRULI": {"token":"18529", "sec":"FINANCE"},
    "CHOLAFIN":   {"token":"685",   "sec":"FINANCE"},
    "MUTHOOTFIN": {"token":"13923", "sec":"FINANCE"},
    "PFC":        {"token":"14299", "sec":"FINANCE"},
    "RECLTD":     {"token":"21614", "sec":"FINANCE"},
    "SHRIRAMFIN": {"token":"4306",  "sec":"FINANCE"},
    "INFY":       {"token":"1594",  "sec":"IT"},
    "TCS":        {"token":"11536", "sec":"IT"},
    "WIPRO":      {"token":"3787",  "sec":"IT"},
    "HCLTECH":    {"token":"7229",  "sec":"IT"},
    "TECHM":      {"token":"13538", "sec":"IT"},
    "LTIM":       {"token":"17818", "sec":"IT"},
    "MPHASIS":    {"token":"4503",  "sec":"IT"},
    "PERSISTENT": {"token":"18365", "sec":"IT"},
    "COFORGE":    {"token":"11543", "sec":"IT"},
    "OFSS":       {"token":"10738", "sec":"IT"},
    "MARUTI":     {"token":"10999", "sec":"AUTO"},
    "MM":         {"token":"2031",  "sec":"AUTO"},
    "TATAMOTORS": {"token":"3456",  "sec":"AUTO"},
    "BAJAJ-AUTO": {"token":"16669", "sec":"AUTO"},
    "HEROMOTOCO": {"token":"1348",  "sec":"AUTO"},
    "EICHERMOT":  {"token":"910",   "sec":"AUTO"},
    "TVSMOTOR":   {"token":"3518",  "sec":"AUTO"},
    "ASHOKLEY":   {"token":"212",   "sec":"AUTO"},
    "SUNPHARMA":  {"token":"3351",  "sec":"PHARMA"},
    "DRREDDY":    {"token":"881",   "sec":"PHARMA"},
    "CIPLA":      {"token":"694",   "sec":"PHARMA"},
    "DIVISLAB":   {"token":"10940", "sec":"PHARMA"},
    "AUROPHARMA": {"token":"275",   "sec":"PHARMA"},
    "ALKEM":      {"token":"13634", "sec":"PHARMA"},
    "BIOCON":     {"token":"524",   "sec":"PHARMA"},
    "UPL":        {"token":"11287", "sec":"PHARMA"},
    "TATASTEEL":  {"token":"3499",  "sec":"METAL"},
    "JSWSTEEL":   {"token":"11723", "sec":"METAL"},
    "HINDALCO":   {"token":"1363",  "sec":"METAL"},
    "COALINDIA":  {"token":"20374", "sec":"METAL"},
    "VEDL":       {"token":"3063",  "sec":"METAL"},
    "NMDC":       {"token":"15332", "sec":"METAL"},
    "HINDUNILVR": {"token":"1394",  "sec":"FMCG"},
    "ITC":        {"token":"1660",  "sec":"FMCG"},
    "BRITANNIA":  {"token":"547",   "sec":"FMCG"},
    "NESTLEIND":  {"token":"17963", "sec":"FMCG"},
    "DABUR":      {"token":"772",   "sec":"FMCG"},
    "GODREJCP":   {"token":"10099", "sec":"FMCG"},
    "MARICO":     {"token":"4067",  "sec":"FMCG"},
    "COLPAL":     {"token":"1367",  "sec":"FMCG"},
    "RELIANCE":   {"token":"2885",  "sec":"ENERGY"},
    "ONGC":       {"token":"2475",  "sec":"ENERGY"},
    "BPCL":       {"token":"526",   "sec":"ENERGY"},
    "NTPC":       {"token":"11630", "sec":"ENERGY"},
    "POWERGRID":  {"token":"14977", "sec":"ENERGY"},
    "TATAPOWER":  {"token":"3426",  "sec":"ENERGY"},
    "IOC":        {"token":"1624",  "sec":"ENERGY"},
    "DLF":        {"token":"14732", "sec":"REALTY"},
    "GODREJPROP": {"token":"9742",  "sec":"REALTY"},
    "OBEROIRLTY": {"token":"20316", "sec":"REALTY"},
    "LT":         {"token":"11483", "sec":"INFRA"},
    "ADANIPORTS": {"token":"15083", "sec":"INFRA"},
    "SIEMENS":    {"token":"3150",  "sec":"INFRA"},
    "ABB":        {"token":"13",    "sec":"INFRA"},
    "BEL":        {"token":"383",   "sec":"INFRA"},
    "HAL":        {"token":"10455", "sec":"INFRA"},
    "BHEL":       {"token":"438",   "sec":"INFRA"},
    "TITAN":      {"token":"3506",  "sec":"CONSUMER"},
    "ASIANPAINT": {"token":"236",   "sec":"CONSUMER"},
    "PIDILITIND": {"token":"2664",  "sec":"CONSUMER"},
    "HAVELLS":    {"token":"10350", "sec":"CONSUMER"},
    "VOLTAS":     {"token":"3718",  "sec":"CONSUMER"},
    "DMART":      {"token":"14413", "sec":"CONSUMER"},
    "TRENT":      {"token":"3721",  "sec":"CONSUMER"},
    "ZOMATO":     {"token":"5097",  "sec":"CONSUMER"},
    "NYKAA":      {"token":"21431", "sec":"CONSUMER"},
    "IRCTC":      {"token":"13611", "sec":"CONSUMER"},
    "ULTRACEMCO": {"token":"11532", "sec":"CEMENT"},
    "GRASIM":     {"token":"1232",  "sec":"CEMENT"},
    "AMBUJACEM":  {"token":"1270",  "sec":"CEMENT"},
    "SHREECEM":   {"token":"3103",  "sec":"CEMENT"},
    "BHARTIARTL": {"token":"10604", "sec":"TELECOM"},
    "INDUSTOWER": {"token":"17491", "sec":"TELECOM"},
}

SECTOR_STOCKS = {
    "BANKING":  ["HDFCBANK","ICICIBANK","SBIN","AXISBANK","KOTAKBANK","INDUSINDBK"],
    "FINANCE":  ["BAJFINANCE","BAJAJFINSV","CHOLAFIN","PFC","RECLTD","MUTHOOTFIN"],
    "IT":       ["INFY","TCS","WIPRO","HCLTECH","TECHM","LTIM"],
    "AUTO":     ["MARUTI","MM","TATAMOTORS","BAJAJ-AUTO","HEROMOTOCO"],
    "PHARMA":   ["SUNPHARMA","DRREDDY","CIPLA","DIVISLAB","AUROPHARMA"],
    "METAL":    ["TATASTEEL","JSWSTEEL","HINDALCO","COALINDIA","VEDL"],
    "FMCG":     ["HINDUNILVR","ITC","BRITANNIA","NESTLEIND","DABUR"],
    "ENERGY":   ["RELIANCE","ONGC","BPCL","NTPC","POWERGRID"],
    "REALTY":   ["DLF","GODREJPROP","OBEROIRLTY"],
    "INFRA":    ["LT","ADANIPORTS","BEL","HAL","BHEL"],
    "CONSUMER": ["TITAN","ASIANPAINT","DMART","TRENT","ZOMATO"],
    "CEMENT":   ["ULTRACEMCO","GRASIM","AMBUJACEM"],
    "TELECOM":  ["BHARTIARTL","INDUSTOWER"],
}

INDICES = {"NIFTY50":"26000","FINNIFTY":"26037"}

# ── PHASE DETECTION ───────────────────────────────────────────────────
def get_phase(ist):
    h,m=ist.hour,ist.minute
    if h==8 and m>=45: return "PREMARKET_845"
    if h==9 and m<15:  return "PREMARKET_900"
    if h==9 and m<30:  return "MARKET_OPEN_915"
    if h==9 and m<45:  return "ORB_930"
    if 9<=h<15:        return "ORB_945_PLUS"
    return "CLOSED"

def get_phase_label(p):
    return {
        "PREMARKET_845":  "8:45 AM — Pre-Market Brief",
        "PREMARKET_900":  "9:00 AM — Pre-Open Analysis",
        "MARKET_OPEN_915":"9:15 AM — Market Open Watch",
        "ORB_930":        "9:30 AM — ORB Confirmation",
        "ORB_945_PLUS":   "9:45 AM — FINNIFTY ORB Signal",
        "CLOSED":         "Market Closed — EOD Data",
    }.get(p,"—")

# ── GIFT NIFTY ────────────────────────────────────────────────────────
def fetch_gift_nifty():
    gift={"ltp":0,"chg":0,"bias":"—","source":"—"}
    hdrs={"User-Agent":"Mozilla/5.0","Accept":"application/json"}
    for sym in ["^NSEI","^CNXNIFTY","NIFTYBEES.NS"]:
        try:
            r=requests.get(
                f"https://query1.finance.yahoo.com/v8/finance/chart/{sym}?interval=1d&range=1d",
                headers=hdrs,timeout=10)
            if r.ok:
                result=r.json().get("chart",{}).get("result",[])
                if result:
                    meta=result[0].get("meta",{})
                    p=float(meta.get("regularMarketPrice",0))
                    prev=float(meta.get("previousClose",0) or meta.get("chartPreviousClose",0))
                    if p>0 and prev>0:
                        chg=round((p-prev)/prev*100,2)
                        gift={"ltp":round(p,2),"chg":chg,
                              "bias":"GAP UP ▲ BULLISH" if chg>0.5 else
                                     "SLIGHT GAP UP" if chg>0.2 else
                                     "GAP DOWN ▼ BEARISH" if chg<-0.5 else
                                     "SLIGHT GAP DOWN" if chg<-0.2 else
                                     "FLAT — WAIT",
                              "source":sym}
                        print(f"Gift Nifty: {p} {chg}% — {gift['bias']}")
                        return gift
        except: continue
    print("Gift Nifty: All sources failed")
    return gift

# ── VIX ───────────────────────────────────────────────────────────────
def fetch_vix():
    vix={"ltp":0,"chg":0,"prev":0,"status":"—","source":"—"}
    hdrs={"User-Agent":"Mozilla/5.0","Accept":"application/json"}
    for url in [
        "https://query1.finance.yahoo.com/v8/finance/chart/%5EINDIAVIX?interval=1d&range=2d",
        "https://query2.finance.yahoo.com/v8/finance/chart/%5EINDIAVIX?interval=1d&range=2d",
    ]:
        try:
            r=requests.get(url,headers=hdrs,timeout=8)
            if r.ok:
                result=r.json().get("chart",{}).get("result",[])
                if result:
                    meta=result[0].get("meta",{})
                    p=float(meta.get("regularMarketPrice",0))
                    prev=float(meta.get("previousClose",0))
                    if p>0:
                        vix={"ltp":round(p,2),"prev":round(prev,2),
                             "chg":round(p-prev,2),
                             "status":"<13 SAFE" if p<13 else "13-16 CAUTION" if p<16 else ">16 HIGH RISK",
                             "source":"Yahoo"}
                        print(f"VIX: {p} — {vix['status']}")
                        return vix
        except: continue
    print("VIX: All sources failed")
    return vix

# ── API HELPERS ───────────────────────────────────────────────────────
def api_hdrs(tok=None):
    h={"Content-Type":"application/json","Accept":"application/json",
       "X-UserType":"USER","X-SourceID":"WEB",
       "X-ClientLocalIP":"192.168.1.1","X-ClientPublicIP":"106.193.147.98",
       "X-MACAddress":"fe80::216e:6507:4b90:3719","X-PrivateKey":API_KEY}
    if tok: h["Authorization"]=f"Bearer {tok}"
    return h

def login():
    totp=pyotp.TOTP(TOTP_SECRET).now()
    print(f"TOTP: {totp}")
    r=requests.post(f"{BASE}/rest/auth/angelbroking/user/v1/loginByPassword",
        json={"clientcode":CLIENT_ID,"password":PIN,"totp":totp},
        headers=api_hdrs(),timeout=15)
    d=r.json()
    if d.get("status") and d.get("data",{}).get("jwtToken"):
        print("✅ Login OK")
        return d["data"]["jwtToken"]
    raise Exception(f"Login failed: {d.get('message')}")

def get_quotes(tok,tokens):
    all_data=[]
    for i in range(0,len(tokens),50):
        batch=tokens[i:i+50]
        try:
            r=requests.post(f"{BASE}/rest/secure/angelbroking/market/v1/quote/",
                json={"mode":"FULL","exchangeTokens":{"NSE":batch}},
                headers=api_hdrs(tok),timeout=15)
            d=r.json()
            if d.get("status") and d.get("data"):
                all_data.extend(d["data"].get("fetched",[]))
        except Exception as e: print(f"Quote error: {e}")
        time.sleep(0.3)
    return all_data

def get_candles(tok,sym_tok,interval,fr,to):
    try:
        r=requests.post(f"{BASE}/rest/secure/angelbroking/historical/v1/getCandleData",
            json={"exchange":"NSE","symboltoken":sym_tok,
                  "interval":interval,"fromdate":fr,"todate":to},
            headers=api_hdrs(tok),timeout=15)
        return r.json()
    except: return {}

# ── INDICATORS ────────────────────────────────────────────────────────
def ema(prices,n):
    if not prices: return 0
    if len(prices)<n: return prices[-1]
    k=2/(n+1); e=sum(prices[:n])/n
    for p in prices[n:]: e=p*k+e*(1-k)
    return round(e,2)

def rsi(prices,n=14):
    if len(prices)<n+1: return 50
    g,l=[],[]
    for i in range(1,len(prices)):
        d=prices[i]-prices[i-1]
        g.append(max(d,0)); l.append(max(-d,0))
    ag=sum(g[:n])/n; al=sum(l[:n])/n
    for i in range(n,len(g)):
        ag=(ag*(n-1)+g[i])/n; al=(al*(n-1)+l[i])/n
    return 100 if al==0 else round(100-100/(1+ag/al))

def macd_bull(prices):
    if len(prices)<26: return True
    k12,k26=2/13,2/27; e12=e26=prices[0]
    for p in prices:
        e12=p*k12+e12*(1-k12); e26=p*k26+e26*(1-k26)
    return e12>e26

def st_bull(candles):
    if not candles: return True
    c=candles[-1]
    return float(c[4])>(float(c[2])+float(c[3]))/2

def vol_ratio(vols):
    if len(vols)<2: return 1.0
    avg=sum(vols[:-1])/len(vols[:-1])
    return round(vols[-1]/avg,1) if avg>0 else 1.0

def calc_vwap(candles):
    if not candles: return 0
    cum_tpv=cum_vol=0
    for c in candles[-20:]:
        tp=(float(c[2])+float(c[3])+float(c[4]))/3
        vol=float(c[5])
        cum_tpv+=tp*vol; cum_vol+=vol
    return round(cum_tpv/cum_vol,2) if cum_vol>0 else 0

# ── EXCEL JOURNAL ─────────────────────────────────────────────────────
def update_excel(signals,ist):
    fname="trading_journal.xlsx"
    try:
        wb=openpyxl.load_workbook(fname); ws=wb.active
    except:
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="SUBHA CAPITAL"
        headers=["Date","Stock","Sector","Side","Score","Pre-Mkt Price",
                 "Open","High","Low","VWAP","ORB High","ORB Low","ORB Range",
                 "Entry","SL","T1","T2","Qty","Max Risk ₹","Potential ₹",
                 "Bias","Result","P&L ₹","Notes"]
        hf=PatternFill(start_color="0D1221",end_color="0D1221",fill_type="solid")
        hfont=Font(bold=True,color="F0A500",size=10)
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=1,column=col,value=h)
            cell.fill=hf; cell.font=hfont
            cell.alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[1].height=28
        widths=[12,12,10,8,7,14,10,10,10,10,10,10,10,10,10,10,10,6,10,12,10,10,10,20]
        for i,w in enumerate(widths,1):
            ws.column_dimensions[get_column_letter(i)].width=w

    date_str=ist.strftime("%d-%b-%Y")
    for s in signals:
        isL=s["side"]=="LONG"
        row=[date_str,s["sym"],s["sec"],s["side"],f"{s['score']}/6",
             s["ltp"],s.get("open",""),s.get("high",""),s.get("low",""),s.get("vwap",0),
             s.get("orb_high",0),s.get("orb_low",0),s.get("orb_range",0),
             s["entry"],s["sl"],s["t1"],s["t2"],s["qty"],
             s.get("max_risk",0),s.get("potential_profit",0),
             s.get("bias",""),"","",""]
        rn=ws.max_row+1
        fc="002D1C" if isL else "2D0010"
        rf=PatternFill(start_color=fc,end_color=fc,fill_type="solid")
        for col,val in enumerate(row,1):
            cell=ws.cell(row=rn,column=col,value=val)
            cell.fill=rf; cell.alignment=Alignment(horizontal='center')
            if col==4: cell.font=Font(color="00D4A0" if isL else "FF4560",bold=True)
            elif col==14: cell.font=Font(color="F0A500",bold=True)
            elif col==15: cell.font=Font(color="FF4560",bold=True)
            elif col in [16,17]: cell.font=Font(color="00D4A0",bold=True)
    wb.save(fname)
    print(f"✅ Excel: {ws.max_row-1} signals")

# ── MAIN ──────────────────────────────────────────────────────────────
def main():
    ist=datetime.now(IST)
    phase=get_phase(ist)
    phase_label=get_phase_label(phase)
    print(f"{'='*60}")
    print(f"SUBHA CAPITAL — {ist.strftime('%d %b %Y %H:%M IST')}")
    print(f"Phase: {phase} — {phase_label}")
    print(f"Universe: Nifty 100 — {len(STOCKS)} stocks")
    print(f"{'='*60}")

    mkt_open=(ist.hour>9 or (ist.hour==9 and ist.minute>=15)) and \
             (ist.hour<15 or (ist.hour==15 and ist.minute<30))

    # ── GIFT NIFTY + VIX ─────────────────────────────────────────────
    print("\nFetching Gift Nifty...")
    gift=fetch_gift_nifty()
    print("Fetching India VIX...")
    vix_data=fetch_vix()

    # ── LOGIN ─────────────────────────────────────────────────────────
    jwt=login()

    # ── QUOTES ───────────────────────────────────────────────────────
    stk_toks=[v["token"] for v in STOCKS.values()]
    idx_toks=list(INDICES.values())
    fetched=get_quotes(jwt,stk_toks+idx_toks)
    Q={}
    for q in fetched:
        t=q.get("symbolToken","")
        for sym,info in STOCKS.items():
            if info["token"]==t: Q[sym]=q; break
        for idx,tok in INDICES.items():
            if tok==t: Q[idx]=q; break
    print(f"Quotes: {len(Q)}/{len(STOCKS)+len(INDICES)}")

    # ── PRICE FUNCTIONS ───────────────────────────────────────────────
    def F(s,field,default=0):
        return float(Q.get(s,{}).get(field,default) or default)

    def ltp(s):
        v=F(s,"ltp")
        return round(v,2) if v>0 else round(F(s,"close"),2)

    def prev_close(s):
        return round(F(s,"close"),2)

    def pct_chg(s):
        pc=F(s,"percentChange")
        if pc!=0: return round(pc,2)
        cur=ltp(s); prev=prev_close(s)
        if cur>0 and prev>0 and cur!=prev:
            return round((cur-prev)/prev*100,2)
        return 0.0

    def pts_chg(s):
        nc=F(s,"netChange")
        if nc!=0: return round(nc,2)
        return round(ltp(s)-prev_close(s),2)

    def day_high(s,c=[]):
        v=round(F(s,"high"),2)
        if v>0: return v
        return round(float(c[-1][2]),2) if c else ltp(s)

    def day_low(s,c=[]):
        v=round(F(s,"low"),2)
        if v>0: return v
        return round(float(c[-1][3]),2) if c else ltp(s)

    def day_open(s,c=[]):
        v=round(F(s,"open"),2)
        if v>0: return v
        return round(float(c[-1][1]),2) if c else ltp(s)

    # ── INDEX DATA ────────────────────────────────────────────────────
    nifty_ltp  = ltp("NIFTY50")
    nifty_chg  = pct_chg("NIFTY50")
    nifty_pts  = pts_chg("NIFTY50")
    nifty_prev = prev_close("NIFTY50")
    fn_ltp     = ltp("FINNIFTY")
    fn_chg     = pct_chg("FINNIFTY")
    fn_pts     = pts_chg("FINNIFTY")
    fn_prev    = prev_close("FINNIFTY")
    print(f"Nifty: {nifty_ltp} | {nifty_chg}% | {nifty_pts}pts")
    print(f"FINNIFTY: {fn_ltp} | {fn_chg}% | {fn_pts}pts")

    bias="BULLISH" if nifty_chg>0.3 else "BEARISH" if nifty_chg<-0.3 else "NEUTRAL"

    # ── FINNIFTY ORB ──────────────────────────────────────────────────
    fn_atm=round(fn_ltp/50)*50 if fn_ltp>0 else 0
    fn_orb={"open":0,"high":0,"low":0,"close":0,"signal":"WAIT","ce_prem":0,"pe_prem":0}

    if phase in ["ORB_945_PLUS","CLOSED","ORB_930","MARKET_OPEN_915"]:
        try:
            fn_date=ist.strftime("%Y-%m-%d")
            for interval,label in [("THIRTY_MINUTE","30min"),("FIFTEEN_MINUTE","15min"),("ONE_MINUTE","1min")]:
                try:
                    cd=get_candles(jwt,"26037",interval,f"{fn_date} 09:00",f"{fn_date} 15:30")
                    cl=cd.get("data",[]) if cd.get("status") else []
                    print(f"FINNIFTY {label}: {len(cl)} candles")
                    if cl:
                        if interval=="THIRTY_MINUTE":
                            c=cl[0]
                            fn_orb["open"]=round(float(c[1]),2)
                            fn_orb["high"]=round(float(c[2]),2)
                            fn_orb["low"]=round(float(c[3]),2)
                            fn_orb["close"]=round(float(c[4]),2)
                        elif interval=="FIFTEEN_MINUTE":
                            c2=cl[:2]
                            fn_orb["open"]=round(float(c2[0][1]),2)
                            fn_orb["high"]=round(max(float(x[2]) for x in c2),2)
                            fn_orb["low"]=round(min(float(x[3]) for x in c2),2)
                            fn_orb["close"]=round(float(c2[-1][4]),2)
                        elif interval=="ONE_MINUTE":
                            c30=cl[:30]
                            fn_orb["open"]=round(float(c30[0][1]),2)
                            fn_orb["high"]=round(max(float(x[2]) for x in c30),2)
                            fn_orb["low"]=round(min(float(x[3]) for x in c30),2)
                            fn_orb["close"]=round(float(c30[-1][4]),2)
                        if fn_orb["high"]>0:
                            if fn_ltp>fn_orb["high"]: fn_orb["signal"]="CE"
                            elif fn_ltp<fn_orb["low"]: fn_orb["signal"]="PE"
                            else: fn_orb["signal"]="MONITOR"
                            print(f"ORB OK: H={fn_orb['high']} L={fn_orb['low']} Sig={fn_orb['signal']}")
                            break
                except Exception as e:
                    print(f"FINNIFTY {label} error: {e}")
        except Exception as e:
            print(f"FINNIFTY ORB error: {e}")

    if fn_ltp>0:
        fn_orb["ce_prem"]=round(fn_ltp*0.005,0)
        fn_orb["pe_prem"]=round(fn_ltp*0.0046,0)

    # ── SECTORS ───────────────────────────────────────────────────────
    sectors=[]
    for sec,stks in SECTOR_STOCKS.items():
        valid=[s for s in stks if ltp(s)>0]
        if not valid: valid=stks[:1]
        chgs=[pct_chg(s) for s in valid]
        sec_chg=round(sum(chgs)/len(chgs),2) if chgs else 0.0
        ldr=max(valid,key=lambda s:abs(pct_chg(s))) if valid else stks[0]
        sectors.append({
            "name":sec,"chg":sec_chg,"leader":ldr,
            "leader_price":ltp(ldr),"leader_chg":pct_chg(ldr),
            "stocks":[{"sym":s,"ltp":ltp(s),"chg":pct_chg(s)} for s in valid]
        })
        print(f"  Sector {sec}: {sec_chg}% | Leader: {ldr} {pct_chg(ldr)}%")
    sectors.sort(key=lambda x:x["chg"],reverse=True)
    print(f"Top: {sectors[0]['name']} {sectors[0]['chg']}% | Bot: {sectors[-1]['name']} {sectors[-1]['chg']}%")

    # ── CANDLES (DAILY) ───────────────────────────────────────────────
    to_d=ist.strftime("%Y-%m-%d %H:%M")
    fr_d=(ist-timedelta(days=60)).strftime("%Y-%m-%d %H:%M")
    candles={}
    print(f"\nFetching daily candles for {len(STOCKS)} stocks...")
    for i,(sym,info) in enumerate(STOCKS.items()):
        try:
            cd=get_candles(jwt,info["token"],"ONE_DAY",fr_d,to_d)
            if cd.get("status") and cd.get("data"):
                candles[sym]=cd["data"]
            time.sleep(0.2)
        except: pass
        if (i+1)%25==0: print(f"  Candles: {i+1}/{len(STOCKS)}")
    print(f"Daily candles: {len(candles)}/{len(STOCKS)}")

    # ── 5-MIN CANDLES FOR ORB CONFIRMATION ───────────────────────────
    # Fetch only during/after market hours for top candidates
    five_min_candles={}
    today=ist.strftime("%Y-%m-%d")
    if phase in ["ORB_930","ORB_945_PLUS","CLOSED"]:
        print(f"\nFetching 5-min candles for ORB confirmation...")
        for i,(sym,info) in enumerate(STOCKS.items()):
            try:
                cd=get_candles(jwt,info["token"],"FIVE_MINUTE",
                               f"{today} 09:15",f"{today} 10:00")
                if cd.get("status") and cd.get("data"):
                    five_min_candles[sym]=cd["data"]
                time.sleep(0.2)
            except: pass
            if (i+1)%25==0: print(f"  5-min: {i+1}/{len(STOCKS)}")
        print(f"5-min candles: {len(five_min_candles)}/{len(STOCKS)}")

    # ── ROCKERS SCAN — ORIGINAL 6 CONDITIONS + 5-MIN CONFIRMATION ────
    top4=[s["name"] for s in sectors[:4]]
    bot4=[s["name"] for s in sectors[-4:]]
    risk_amt=CAPITAL*RISK_PCT/100
    longs=[]; shorts=[]; skipped=[]

    def get_5min_confirmation(sym, orb_high, orb_low, direction):
        """
        5-min candle confirmation after 9:30 AM ORB
        
        LONG:  9:35 AM candle closes ABOVE ORB High = confirmed
        SHORT: 9:35 AM candle closes BELOW ORB Low  = confirmed
        
        Returns: "CONFIRMED", "REJECTED", "PENDING" (no data yet)
        """
        fm=five_min_candles.get(sym,[])
        if not fm:
            return "PENDING"  # No 5-min data = pre-market or data unavailable

        # Find the 9:35 AM candle (2nd 5-min candle of the day)
        # Index 0 = 9:15-9:20, Index 1 = 9:20-9:25, Index 2 = 9:25-9:30
        # Index 3 = 9:30-9:35 ← ORB candle
        # Index 4 = 9:35-9:40 ← confirmation candle
        confirm_candle = fm[4] if len(fm)>4 else fm[-1] if fm else None

        if not confirm_candle:
            return "PENDING"

        c_close = float(confirm_candle[4])
        c_high  = float(confirm_candle[2])
        c_low   = float(confirm_candle[3])

        if direction=="LONG":
            if c_close > orb_high:
                return "CONFIRMED"  # 9:35 candle closed above ORB high ✅
            elif c_close < orb_low:
                return "REJECTED"   # 9:35 candle closed below ORB low ❌
            else:
                return "PENDING"    # Still inside ORB range — wait

        elif direction=="SHORT":
            if c_close < orb_low:
                return "CONFIRMED"  # 9:35 candle closed below ORB low ✅
            elif c_close > orb_high:
                return "REJECTED"   # 9:35 candle closed above ORB high ❌
            else:
                return "PENDING"    # Still inside ORB range — wait

        return "PENDING"

    for sym,info in STOCKS.items():
        p=ltp(sym)
        if p<=0: continue

        c=candles.get(sym,[])
        closes=[float(x[4]) for x in c] if c else [p]
        vols=[float(x[5]) for x in c] if c else [1]

        h=day_high(sym,c); l=day_low(sym,c); op=day_open(sym,c)
        if h<=0: h=p
        if l<=0: l=p
        if op<=0: op=p

        vwap_val=calc_vwap(c)
        e20=ema(closes,20); r=rsi(closes); vr=vol_ratio(vols)
        macd_b=macd_bull(closes); st_b=st_bull(c)

        # ── 2 SIMPLE SMART FILTERS ONLY ───────────────────────────────
        prev=prev_close(sym)
        gap_pct=abs(op-prev)/prev*100 if prev>0 else 0
        first_candle_pct=abs(h-op)/op*100 if op>0 else 0
        skip_reason=None

        if first_candle_pct>2.0:
            skip_reason=f"First candle {first_candle_pct:.1f}% — already moved"
        elif gap_pct>5.0:
            skip_reason=f"Gap {gap_pct:.1f}% at open — results day, priced in"

        if skip_reason:
            skipped.append({
                "sym":sym,"sec":info["sec"],"ltp":p,
                "chg":pct_chg(sym),"reason":skip_reason
            })
            continue

        # ── ORIGINAL 6-CONDITION ROCKERS ──────────────────────────────
        if info["sec"] in top4:
            vwap_ok=p>vwap_val if vwap_val>0 else p>e20
            conds={
                "EMA20":  p>e20,
                "SUPERT": st_b,
                "RSI":    50<=r<=70,
                "MACD":   macd_b,
                "VOL":    vr>=1.5,
                "VWAP":   vwap_ok,
            }
            sc=sum(conds.values())
            orb_range=round(h-l,2)
            orb_mid=round((h+l)/2,2)
            en=round(h+0.05,2)
            sl=round(orb_mid-0.05,2)
            rp=round(en-sl,2)
            if rp<=0: continue
            t1=round(en+orb_range*0.5,2)
            t2=round(en+orb_range*1.0,2)
            qty=min(int((CAPITAL*MARGIN)/en),max(1,int(risk_amt/rp)*3))

            # 5-min confirmation
            confirm=get_5min_confirmation(sym,h,l,"LONG")

            # Skip if 5-min candle REJECTED the long signal
            if confirm=="REJECTED":
                skipped.append({
                    "sym":sym,"sec":info["sec"],"ltp":p,
                    "chg":pct_chg(sym),
                    "reason":f"5-min candle rejected LONG — closed below ORB low"
                })
                print(f"  SKIP {sym}: 5-min rejected LONG")
                continue

            longs.append({
                "sym":sym,"sec":info["sec"],"ltp":p,"chg":pct_chg(sym),
                "open":op,"high":h,"low":l,"vwap":vwap_val,
                "orb_high":h,"orb_low":l,"orb_mid":orb_mid,"orb_range":orb_range,
                "score":sc,"rsi":r,"vol_ratio":vr,
                "conds":{k:bool(v) for k,v in conds.items()},
                "entry":en,"sl":sl,"t1":t1,"t2":t2,
                "gtt_trigger":en,"gtt_limit":round(en+0.05,2),
                "sl_trigger":sl,"sl_limit":round(sl-0.05,2),
                "qty":qty,"risk_per_share":rp,
                "max_risk":round(rp*qty,0),
                "potential_profit":round((t1-en)*qty,0),
                "confirmation":confirm,
                "side":"LONG","bias":bias,"phase":phase
            })

        if info["sec"] in bot4:
            vwap_ok=p<vwap_val if vwap_val>0 else p<e20
            conds={
                "EMA20":  p<e20,
                "SUPERT": not st_b,
                "RSI":    30<=r<=50,
                "MACD":   not macd_b,
                "VOL":    vr>=1.5,
                "VWAP":   vwap_ok,
            }
            sc=sum(conds.values())
            orb_range=round(h-l,2)
            orb_mid=round((h+l)/2,2)
            en=round(l-0.05,2)
            sl=round(orb_mid+0.05,2)
            rp=round(sl-en,2)
            if rp<=0: continue
            t1=round(en-orb_range*0.5,2)
            t2=round(en-orb_range*1.0,2)
            qty=min(int((CAPITAL*MARGIN)/en),max(1,int(risk_amt/rp)*3))

            # 5-min confirmation
            confirm=get_5min_confirmation(sym,h,l,"SHORT")

            # Skip if 5-min candle REJECTED the short signal
            if confirm=="REJECTED":
                skipped.append({
                    "sym":sym,"sec":info["sec"],"ltp":p,
                    "chg":pct_chg(sym),
                    "reason":f"5-min candle rejected SHORT — closed above ORB high"
                })
                print(f"  SKIP {sym}: 5-min rejected SHORT")
                continue

            shorts.append({
                "sym":sym,"sec":info["sec"],"ltp":p,"chg":pct_chg(sym),
                "open":op,"high":h,"low":l,"vwap":vwap_val,
                "orb_high":h,"orb_low":l,"orb_mid":orb_mid,"orb_range":orb_range,
                "score":sc,"rsi":r,"vol_ratio":vr,
                "conds":{k:bool(v) for k,v in conds.items()},
                "entry":en,"sl":sl,"t1":t1,"t2":t2,
                "gtt_trigger":en,"gtt_limit":round(en-0.05,2),
                "sl_trigger":sl,"sl_limit":round(sl+0.05,2),
                "qty":qty,"risk_per_share":rp,
                "max_risk":round(rp*qty,0),
                "potential_profit":round((en-t1)*qty,0),
                "confirmation":confirm,
                "side":"SHORT","bias":bias,"phase":phase
            })

    longs.sort(key=lambda x:x["score"],reverse=True)
    shorts.sort(key=lambda x:x["score"],reverse=True)
    print(f"\nLongs: {len(longs)} | Shorts: {len(shorts)} | Skipped: {len(skipped)}")
    print(f"Top 2 Long:  {[s['sym'] for s in longs[:2]]}")
    print(f"Top 2 Short: {[s['sym'] for s in shorts[:2]]}")
    print(f"Skipped: {[s['sym'] for s in skipped]}")

    # ── EXCEL ─────────────────────────────────────────────────────────
    trade_signals=[{**s} for s in longs[:2]+shorts[:2]]
    if trade_signals:
        try: update_excel(trade_signals,ist)
        except Exception as e: print(f"Excel error: {e}")

    # ── SAVE DATA.JSON ────────────────────────────────────────────────
    data={
        "generated_at":  ist.strftime("%d %b %Y %I:%M %p IST"),
        "phase":         phase,
        "phase_label":   phase_label,
        "market_status": "OPEN" if mkt_open else "CLOSED",
        "market":{
            "nifty50":{"ltp":nifty_ltp,"chg":nifty_chg,"pts":nifty_pts,"prev":nifty_prev},
            "vix":{"ltp":vix_data["ltp"],"chg":vix_data["chg"],"prev":vix_data["prev"],
                   "status":vix_data["status"],"source":vix_data["source"]},
            "bias":bias
        },
        "gift_nifty": gift,
        "pre_market":{
            "analysis":f"Gift Nifty {gift['bias']}. {bias} bias. VIX {vix_data['ltp'] or 'N/A'}.",
            "top_gainers":sorted([{"sym":s,"ltp":ltp(s),"chg":pct_chg(s),"sec":STOCKS[s]["sec"]}
                                  for s in STOCKS if ltp(s)>0],
                                 key=lambda x:x["chg"],reverse=True)[:3],
            "top_losers": sorted([{"sym":s,"ltp":ltp(s),"chg":pct_chg(s),"sec":STOCKS[s]["sec"]}
                                  for s in STOCKS if ltp(s)>0],
                                 key=lambda x:x["chg"])[:3],
        },
        "finnifty":{
            "ltp":fn_ltp,"chg":fn_chg,"pts":fn_pts,"prev":fn_prev,"atm":fn_atm,
            "orb_open":fn_orb["open"],"orb_high":fn_orb["high"],
            "orb_low":fn_orb["low"],"orb_close":fn_orb["close"],
            "signal":fn_orb["signal"],
            "ce_premium":fn_orb["ce_prem"],"pe_premium":fn_orb["pe_prem"],
            "ce_sl":round(fn_orb["ce_prem"]*0.7,0),
            "ce_tgt":round(fn_orb["ce_prem"]*1.5,0),
            "pe_sl":round(fn_orb["pe_prem"]*0.7,0),
            "pe_tgt":round(fn_orb["pe_prem"]*1.5,0),
        },
        "sectors":    sectors,
        "long":       longs[:5],
        "short":      shorts[:5],
        "trade_long":  longs[:2],
        "trade_short": shorts[:2],
        "skipped":    skipped[:10],
        "capital":    CAPITAL,
        "risk_pct":   RISK_PCT,
        "risk_amt":   risk_amt,
        "universe":   f"Nifty 100 — {len(STOCKS)} stocks",
    }

    with open("data.json","w") as f:
        json.dump(data,f,indent=2)
    print(f"\n✅ data.json saved!")
    print(f"Trade: {longs[0]['sym'] if longs else 'No long'} LONG + {shorts[0]['sym'] if shorts else 'No short'} SHORT")
    print(f"{'='*60}")

if __name__=="__main__":
    main()
